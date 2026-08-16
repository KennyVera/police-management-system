"""Reportes de zona (PDF / Excel) a partir de ClickHouse, con aislamiento geográfico."""

from __future__ import annotations

import io
import math
import urllib.request
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date, datetime, timedelta
from typing import Any

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    HRFlowable,
    Image as RLImage,
    KeepTogether,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from reportlab.graphics.shapes import Circle, Drawing, Line, Polygon, Rect, String
from reportlab.graphics.charts.barcharts import HorizontalBarChart, VerticalBarChart
from reportlab.graphics.charts.lineplots import LinePlot
from reportlab.graphics.widgets.markers import makeMarker

from tactico.services.clickhouse_client import execute_readonly
from tactico.services.geo_scope import ZoneScope

FACT = "police_analytics.fact_partes_policiales"


def _dt_bounds(fecha_desde: date | None, fecha_hasta: date | None):
    hasta = fecha_hasta or date.today()
    desde = fecha_desde or (hasta - timedelta(days=30))
    return (
        desde,
        hasta,
        datetime.combine(desde, datetime.min.time()),
        datetime.combine(hasta, datetime.max.time().replace(microsecond=0)),
    )


def collect_zone_report_data(
    scope: ZoneScope,
    *,
    fecha_desde: date | None = None,
    fecha_hasta: date | None = None,
    audiencia: str = "ALTO_MANDO",
) -> dict[str, Any]:
    """Agrega KPIs tácticos de la zona desde ClickHouse (solo lectura)."""
    desde, hasta, dt_desde, dt_hasta = _dt_bounds(fecha_desde, fecha_hasta)
    params = {
        **scope.geo_params,
        "fecha_desde": dt_desde,
        "fecha_hasta": dt_hasta,
    }
    date_clause = (
        " AND fecha_hora >= {fecha_desde:DateTime}"
        " AND fecha_hora <= {fecha_hasta:DateTime}"
    )

    stats_sql = (
        "SELECT "
        "countIf(toYYYYMM(fecha_hora) = toYYYYMM(today())) AS mes_actual, "
        "countIf(toYYYYMM(fecha_hora) = toYYYYMM(addMonths(today(), -1))) AS mes_anterior, "
        "toUInt32(count()) AS total_rango "
        f"FROM {FACT} WHERE 1 = 1 " + scope.geo_sql + date_clause
    )
    tipo_sql = (
        "SELECT tipo_delito, toUInt32(count()) AS total "
        f"FROM {FACT} WHERE 1 = 1 " + scope.geo_sql + date_clause
        + " GROUP BY tipo_delito ORDER BY total DESC LIMIT 25"
    )
    dist_sql = (
        "SELECT sector_zona AS distrito, toUInt32(count()) AS total_partes, "
        "toUInt32(countIf(upper(prioridad) IN ('ALTA','CRITICA','ALTO','CRITICO'))) AS criticos "
        f"FROM {FACT} WHERE 1 = 1 " + scope.geo_sql + date_clause
        + " AND sector_zona != '' GROUP BY sector_zona ORDER BY total_partes DESC LIMIT 30"
    )

    stats = execute_readonly(stats_sql, params)
    por_tipo = execute_readonly(tipo_sql, params)
    por_distrito = execute_readonly(dist_sql, params)
    row = stats[0] if stats else {}

    return {
        "audiencia": audiencia,
        "generado_en": datetime.now().isoformat(timespec="seconds"),
        "jurisdiccion": {
            "id": scope.jurisdiccion_id,
            "nombre": scope.jurisdiccion_nombre,
            "codigo": scope.jurisdiccion_codigo,
        },
        "periodo": {
            "fecha_desde": desde.isoformat(),
            "fecha_hasta": hasta.isoformat(),
        },
        "resumen": {
            "total_rango": int(row.get("total_rango") or 0),
            "mes_actual": int(row.get("mes_actual") or 0),
            "mes_anterior": int(row.get("mes_anterior") or 0),
        },
        "por_tipo": [
            {
                "tipo_delito": r.get("tipo_delito") or "Sin clasificar",
                "total": int(r.get("total") or 0),
            }
            for r in por_tipo
        ],
        "por_distrito": [
            {
                "distrito": r.get("distrito") or "",
                "total_partes": int(r.get("total_partes") or 0),
                "criticos": int(r.get("criticos") or 0),
            }
            for r in por_distrito
        ],
    }


def build_zone_report_pdf(data: dict[str, Any], *, titulo: str, emisor: str) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=1.6 * cm,
        rightMargin=1.6 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "TitleZona",
        parent=styles["Heading1"],
        fontSize=16,
        textColor=colors.HexColor("#2f4d8a"),
        spaceAfter=8,
    )
    body = []
    audiencia = data.get("audiencia") or "ALTO_MANDO"
    audiencia_label = (
        "Alto Mando / Visor Ejecutivo"
        if audiencia == "ALTO_MANDO"
        else "Autoridades civiles (Gobernación / Alcaldía / Prefectura)"
    )
    jur = data.get("jurisdiccion") or {}
    periodo = data.get("periodo") or {}
    resumen = data.get("resumen") or {}

    body.append(Paragraph(titulo, title_style))
    body.append(Paragraph(f"<b>Audiencia:</b> {audiencia_label}", styles["Normal"]))
    body.append(
        Paragraph(
            f"<b>Jurisdicción:</b> {jur.get('nombre') or '—'} ({jur.get('codigo') or '—'})",
            styles["Normal"],
        )
    )
    body.append(
        Paragraph(
            f"<b>Periodo:</b> {periodo.get('fecha_desde')} → {periodo.get('fecha_hasta')}",
            styles["Normal"],
        )
    )
    body.append(Paragraph(f"<b>Elaborado por:</b> {emisor}", styles["Normal"]))
    body.append(Paragraph(f"<b>Generado:</b> {data.get('generado_en')}", styles["Normal"]))
    body.append(Spacer(1, 0.4 * cm))
    body.append(Paragraph("Resumen operativo", styles["Heading2"]))
    body.append(
        Paragraph(
            f"Total de partes en el periodo: <b>{resumen.get('total_rango', 0)}</b><br/>"
            f"Mes calendario actual: <b>{resumen.get('mes_actual', 0)}</b> · "
            f"Mes anterior: <b>{resumen.get('mes_anterior', 0)}</b>",
            styles["Normal"],
        )
    )
    body.append(Spacer(1, 0.35 * cm))

    body.append(Paragraph("Delitos por tipo", styles["Heading2"]))
    tipo_rows = [["Tipo de delito", "Total"]] + [
        [r["tipo_delito"], str(r["total"])] for r in data.get("por_tipo") or []
    ]
    if len(tipo_rows) == 1:
        tipo_rows.append(["Sin datos", "0"])
    t1 = Table(tipo_rows, colWidths=[12 * cm, 3 * cm])
    t1.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2f4d8a")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#d1d5db")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("PADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    body.append(t1)
    body.append(Spacer(1, 0.4 * cm))

    body.append(Paragraph("Ranking por distrito / sector", styles["Heading2"]))
    dist_rows = [["Distrito", "Partes", "Críticos"]] + [
        [r["distrito"], str(r["total_partes"]), str(r["criticos"])]
        for r in data.get("por_distrito") or []
    ]
    if len(dist_rows) == 1:
        dist_rows.append(["Sin datos", "0", "0"])
    t2 = Table(dist_rows, colWidths=[9 * cm, 3 * cm, 3 * cm])
    t2.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e3a5f")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#d1d5db")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("PADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    body.append(t2)
    body.append(Spacer(1, 0.5 * cm))
    body.append(
        Paragraph(
            "<i>Documento generado automáticamente por el Módulo de Inteligencia Táctica. "
            "Los datos están filtrados exclusivamente a la jurisdicción del emisor "
            "(aislamiento geográfico).</i>",
            styles["Normal"],
        )
    )
    doc.build(body)
    return buffer.getvalue()


def build_zone_report_excel(data: dict[str, Any]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"
    jur = data.get("jurisdiccion") or {}
    periodo = data.get("periodo") or {}
    resumen = data.get("resumen") or {}
    ws.append(["Informe de seguridad de zona"])
    ws.append(["Jurisdicción", jur.get("nombre") or ""])
    ws.append(["Código", jur.get("codigo") or ""])
    ws.append(["Audiencia", data.get("audiencia") or ""])
    ws.append(["Desde", periodo.get("fecha_desde") or ""])
    ws.append(["Hasta", periodo.get("fecha_hasta") or ""])
    ws.append(["Generado", data.get("generado_en") or ""])
    ws.append([])
    ws.append(["Total periodo", resumen.get("total_rango", 0)])
    ws.append(["Mes actual", resumen.get("mes_actual", 0)])
    ws.append(["Mes anterior", resumen.get("mes_anterior", 0)])
    ws["A1"].font = Font(bold=True, size=14, color="2F4D8A")

    ws2 = wb.create_sheet("Por tipo")
    ws2.append(["Tipo de delito", "Total"])
    for r in data.get("por_tipo") or []:
        ws2.append([r["tipo_delito"], r["total"]])

    ws3 = wb.create_sheet("Por distrito")
    ws3.append(["Distrito", "Partes", "Críticos"])
    for r in data.get("por_distrito") or []:
        ws3.append([r["distrito"], r["total_partes"], r["criticos"]])

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ——— Snapshot del Dashboard táctico (misma data que /api/tactico/panel/) ———

PURPLE = colors.HexColor("#7c5cbf")
PURPLE_DARK = colors.HexColor("#5b3d9a")
SLATE = colors.HexColor("#334155")
MUTED = colors.HexColor("#64748b")
SOFT = colors.HexColor("#f5f0ff")
LINE = colors.HexColor("#e2e8f0")
ANALYSIS_BG = colors.HexColor("#f8f5ff")


def _delta_label(value) -> str:
    if value is None:
        return "Sin comparación"
    sign = "+" if value >= 0 else ""
    return f"{sign}{value}% vs periodo anterior"


def _styled_data_table(rows: list[list[str]], col_widths: list) -> Table:
    t = Table(rows, colWidths=col_widths)
    t.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), PURPLE_DARK),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.35, LINE),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, SOFT]),
                ("ALIGN", (1, 0), (-1, -1), "CENTER"),
                ("PADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    return t


def _analysis_box(texto: str, style: ParagraphStyle) -> Table:
    inner = Paragraph(f"<b>Análisis para la decisión:</b> {texto}", style)
    box = Table([[inner]], colWidths=[26 * cm])
    box.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), ANALYSIS_BG),
                ("BOX", (0, 0), (-1, -1), 0.8, PURPLE),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 7),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
            ]
        )
    )
    return box


def _kpi_card_table(rows: list[list[str]]) -> Table:
    t = Table(rows, colWidths=[3.6 * cm] * 5)
    t.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), SOFT),
                ("BOX", (0, 0), (-1, -1), 0.8, PURPLE),
                ("INNERGRID", (0, 0), (-1, -1), 0.4, LINE),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 7.5),
                ("TEXTCOLOR", (0, 0), (-1, 0), MUTED),
                ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 1), (-1, 1), 14),
                ("TEXTCOLOR", (0, 1), (-1, 1), PURPLE_DARK),
                ("FONTSIZE", (0, 2), (-1, -1), 7),
                ("TEXTCOLOR", (0, 2), (-1, -1), SLATE),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    return t


def _tipologia_bar_chart(tipologia: list[dict]) -> Drawing | None:
    data = tipologia[:8]
    if not data:
        return None
    values = [[int(r.get("total") or 0) for r in data]]
    labels = [(r.get("tipo_delito") or "—")[:22] for r in data]
    drawing = Drawing(520, 155)
    chart = HorizontalBarChart()
    chart.x = 140
    chart.y = 15
    chart.height = 125
    chart.width = 360
    chart.data = values
    chart.categoryAxis.categoryNames = labels
    chart.bars[0].fillColor = PURPLE
    chart.valueAxis.valueMin = 0
    chart.categoryAxis.labels.fontSize = 8
    chart.valueAxis.labels.fontSize = 7
    drawing.add(chart)
    return drawing


def _evolucion_line_chart(evolucion: list[dict]) -> Drawing | None:
    pts = evolucion[-30:] if evolucion else []
    if len(pts) < 2:
        return None
    series = [(i, float(r.get("total") or 0)) for i, r in enumerate(pts)]
    drawing = Drawing(520, 150)
    lp = LinePlot()
    lp.x = 40
    lp.y = 28
    lp.height = 105
    lp.width = 460
    lp.data = [series]
    lp.lines[0].strokeColor = PURPLE
    lp.lines[0].strokeWidth = 2
    lp.lines[0].symbol = makeMarker("FilledCircle")
    lp.lines[0].symbol.size = 3.5
    lp.lines[0].symbol.fillColor = PURPLE_DARK
    lp.xValueAxis.valueMin = 0
    lp.xValueAxis.valueMax = max(len(pts) - 1, 1)
    lp.yValueAxis.valueMin = 0
    ymax = max(v for _, v in series) or 1
    lp.yValueAxis.valueMax = ymax * 1.15
    lp.xValueAxis.labels.fontSize = 6
    lp.yValueAxis.labels.fontSize = 7
    drawing.add(lp)
    first = pts[0].get("fecha") or pts[0].get("label") or ""
    last = pts[-1].get("fecha") or pts[-1].get("label") or ""
    drawing.add(String(40, 8, str(first)[:12], fontSize=7, fillColor=MUTED))
    drawing.add(String(430, 8, str(last)[:12], fontSize=7, fillColor=MUTED))
    return drawing


def _ranking_bar_chart(ranking: list[dict]) -> Drawing | None:
    data = ranking[:8]
    if not data:
        return None
    values = [[int(r.get("total") or r.get("incidentes") or 0) for r in data]]
    labels = [(r.get("distrito") or r.get("nombre") or "—")[:18] for r in data]
    drawing = Drawing(520, 155)
    chart = VerticalBarChart()
    chart.x = 40
    chart.y = 35
    chart.height = 105
    chart.width = 460
    chart.data = values
    chart.categoryAxis.categoryNames = labels
    chart.bars[0].fillColor = PURPLE
    chart.valueAxis.valueMin = 0
    chart.categoryAxis.labels.angle = 25
    chart.categoryAxis.labels.fontSize = 7
    chart.categoryAxis.labels.boxAnchor = "ne"
    chart.valueAxis.labels.fontSize = 7
    drawing.add(chart)
    return drawing


def _analisis_kpis(kpis: dict) -> str:
    ind = kpis.get("indice_delictivo") or {}
    ef = kpis.get("efectividad") or {}
    imp = kpis.get("mayor_impacto") or {}
    alerta = kpis.get("alerta_roja") or {}
    fuerza = kpis.get("fuerza_efectiva") or {}
    total = ind.get("total") or 0
    delta = ind.get("delta_pct")
    tendencia = (
        "estable respecto al periodo anterior"
        if delta is None or delta == 0
        else ("al alza" if delta > 0 else "a la baja")
    )
    delito = imp.get("tipo_delito") or "sin tipificar"
    pct = imp.get("pct")
    alerta_n = alerta.get("total") or 0
    detenidos = ef.get("detenidos") or 0
    op = fuerza.get("porcentaje") or 0
    partes = [
        f"Se registraron <b>{total}</b> incidentes en el periodo ({tendencia}).",
        f"El delito de mayor peso relativo es <b>{delito}</b>"
        + (f" ({pct}% del total)" if pct is not None else "")
        + ".",
        f"La efectividad operativa reporta <b>{detenidos}</b> detenidos; "
        f"úselo para contrastar presión delictiva vs. capacidad de captura.",
    ]
    if alerta_n:
        nombres = ", ".join(alerta.get("distritos") or []) or "varios sectores"
        partes.append(
            f"Hay <b>{alerta_n}</b> distrito(s) en alerta roja ({nombres}): priorice refuerzo y supervisión."
        )
    else:
        partes.append(
            "Ningún distrito supera el umbral de alerta roja; mantenga monitoreo preventivo."
        )
    partes.append(
        f"La fuerza efectiva desplegada hoy es del <b>{op}%</b>: "
        "si es baja frente a la carga delictiva, reasigne turnos o escuadras."
    )
    return " ".join(partes)


def _analisis_tipologia(tip: list[dict]) -> str:
    if not tip:
        return (
            "No hay tipología suficiente en el periodo filtrado. "
            "Verifique el rango de fechas o amplíe el filtro de distrito/delito antes de decidir recursos."
        )
    top = tip[0]
    nombre = top.get("tipo_delito") or "Sin clasificar"
    total = top.get("total") or 0
    pct = top.get("pct") or 0
    segundo = tip[1] if len(tip) > 1 else None
    cola = tip[2:5]
    texto = (
        f"Este gráfico ordena los delitos por volumen. <b>{nombre}</b> lidera con "
        f"<b>{total}</b> casos (<b>{pct}%</b> del total tipificado): concentre patrullaje, "
        f"investigaciones y campañas de prevención en esa tipología."
    )
    if segundo:
        texto += (
            f" Le sigue <b>{segundo.get('tipo_delito')}</b> "
            f"({segundo.get('total') or 0} casos, {segundo.get('pct') or 0}%), "
            "útil para un segundo eje operativo."
        )
    if cola:
        nombres = ", ".join(r.get("tipo_delito") or "—" for r in cola)
        texto += (
            f" El resto ({nombres}) aporta volumen menor pero no debe ignorarse si tiene alto impacto social."
        )
    texto += (
        " Interprete barras largas como prioridad de despliegue; barras cortas, como seguimiento selectivo."
    )
    return texto


def _analisis_evolucion(evo: list[dict]) -> str:
    if not evo:
        return (
            "Sin serie diaria disponible. No es posible detectar picos temporales; "
            "ajuste el periodo e intente nuevamente."
        )
    valores = [int(r.get("total") or 0) for r in evo]
    total = sum(valores)
    max_v = max(valores) if valores else 0
    min_v = min(valores) if valores else 0
    idx_max = valores.index(max_v) if valores else 0
    dia_pico = (evo[idx_max].get("fecha") or "—") if evo else "—"
    mitad = len(valores) // 2 or 1
    primera = sum(valores[:mitad]) / mitad
    segunda = sum(valores[mitad:]) / max(len(valores) - mitad, 1)
    if segunda > primera * 1.15:
        tendencia = "tendencia creciente en la segunda mitad del periodo"
    elif segunda < primera * 0.85:
        tendencia = "tendencia decreciente hacia el final del periodo"
    else:
        tendencia = "comportamiento relativamente estable a lo largo del periodo"
    return (
        f"La curva muestra la incidencia diaria (total del periodo: <b>{total}</b>). "
        f"Se observa {tendencia}. El pico máximo fue <b>{max_v}</b> incidentes el "
        f"<b>{dia_pico}</b> (mínimo diario: {min_v}). "
        "Use los picos para programar operativos reforzados esos días/semanas y los valles "
        "para mantenimiento o redistribución de personal."
    )


def _analisis_ranking(ranking: list[dict]) -> str:
    if not ranking:
        return (
            "No hay ranking territorial con los filtros actuales. "
            "Sin este mapa de carga no conviene redistribuir escuadras a ciegas."
        )
    top = ranking[0]
    nombre = top.get("distrito") or top.get("nombre") or "—"
    total = top.get("total") or top.get("incidentes") or 0
    nivel = (top.get("nivel") or "—").capitalize()
    criticos = [r for r in ranking if (r.get("nivel") or "").lower() == "critico"]
    bajos = [r for r in ranking if (r.get("nivel") or "").lower() == "bajo"]
    texto = (
        f"<b>{nombre}</b> concentra la mayor carga con <b>{total}</b> incidentes "
        f"(nivel <b>{nivel}</b>). Priorice supervisión, refuerzo y control de calidad en ese sector."
    )
    if len(ranking) > 1:
        segundo = ranking[1]
        texto += (
            f" El segundo foco es <b>{segundo.get('distrito') or '—'}</b> "
            f"({segundo.get('total') or 0} incidentes)."
        )
    if criticos:
        names = ", ".join(c.get("distrito") or "—" for c in criticos[:4])
        texto += f" Distritos en nivel crítico: <b>{names}</b> — exigen plan inmediato."
    if bajos:
        texto += (
            f" Hay {len(bajos)} sector(es) en nivel bajo: evalúe si puede ceder personal "
            "temporalmente hacia los distritos de mayor presión."
        )
    texto += (
        " Interprete barras altas como demanda operativa; barras bajas como capacidad liberable."
    )
    return texto


def build_dashboard_snapshot_pdf(panel: dict[str, Any], *, emisor: str) -> bytes:
    """PDF del Dashboard táctico: KPIs + gráficos con análisis, sin títulos huérfanos."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=1.4 * cm,
        rightMargin=1.4 * cm,
        topMargin=1.1 * cm,
        bottomMargin=1.1 * cm,
    )
    styles = getSampleStyleSheet()
    title = ParagraphStyle(
        "DzTitle",
        parent=styles["Heading1"],
        fontSize=17,
        textColor=PURPLE_DARK,
        spaceAfter=2,
        leading=20,
    )
    kicker = ParagraphStyle(
        "DzKicker",
        parent=styles["Normal"],
        fontSize=9,
        textColor=PURPLE,
        fontName="Helvetica-Bold",
        spaceAfter=2,
    )
    h2 = ParagraphStyle(
        "DzH2",
        parent=styles["Heading2"],
        fontSize=12,
        textColor=SLATE,
        spaceBefore=0,
        spaceAfter=6,
    )
    normal = ParagraphStyle(
        "DzNormal",
        parent=styles["Normal"],
        fontSize=9,
        textColor=SLATE,
        leading=12,
    )
    analysis_style = ParagraphStyle(
        "DzAnalysis",
        parent=styles["Normal"],
        fontSize=8.5,
        textColor=SLATE,
        leading=11.5,
    )
    small = ParagraphStyle(
        "DzSmall",
        parent=styles["Normal"],
        fontSize=8,
        textColor=MUTED,
        leading=10,
    )

    jur = panel.get("jurisdiccion") or {}
    filtros = panel.get("filtros") or {}
    kpis = panel.get("kpis") or {}
    tip = panel.get("tipologia") or []
    evo = panel.get("evolucion") or []
    ranking = panel.get("ranking_barras") or []
    resumen = panel.get("resumen_ejecutivo") or ""
    generado = datetime.now().strftime("%Y-%m-%d %H:%M")

    ind = kpis.get("indice_delictivo") or {}
    ef = kpis.get("efectividad") or {}
    imp = kpis.get("mayor_impacto") or {}
    alerta = kpis.get("alerta_roja") or {}
    fuerza = kpis.get("fuerza_efectiva") or {}

    body: list[Any] = []

    # —— Portada / KPIs ——
    header_block = [
        Paragraph("INTELIGENCIA TÁCTICA", kicker),
        Paragraph(
            f"Dashboard de zona — {jur.get('nombre') or 'Jurisdicción'}",
            title,
        ),
        Paragraph(
            f"<b>Periodo:</b> {filtros.get('fecha_desde') or '—'} → {filtros.get('fecha_hasta') or '—'} "
            f"&nbsp;&nbsp;|&nbsp;&nbsp; <b>Distrito:</b> {filtros.get('distrito') or 'Todos'} "
            f"&nbsp;&nbsp;|&nbsp;&nbsp; <b>Delito:</b> {filtros.get('tipo_delito') or 'Todos'}",
            normal,
        ),
        Paragraph(
            f"Elaborado por <b>{emisor}</b> · Generado {generado} · Código {jur.get('codigo') or '—'}",
            small,
        ),
        Spacer(1, 0.12 * cm),
        HRFlowable(width="100%", thickness=1.2, color=PURPLE, spaceAfter=8),
        Paragraph("Indicadores clave (KPIs)", h2),
        _kpi_card_table(
            [
                [
                    "ÍNDICE DELICTIVO",
                    "EFECTIVIDAD",
                    "MAYOR IMPACTO",
                    "ALERTA ROJA",
                    "FUERZA EFECTIVA",
                ],
                [
                    str(ind.get("total") or 0),
                    str(ef.get("detenidos") or 0),
                    str(imp.get("tipo_delito") or "—")[:16],
                    str(alerta.get("total") or 0),
                    f"{fuerza.get('porcentaje') or 0}%",
                ],
                [
                    _delta_label(ind.get("delta_pct")),
                    _delta_label(ef.get("delta_pct")),
                    _delta_label(imp.get("delta_pct")),
                    f"{len(alerta.get('distritos') or [])} distritos",
                    f"{fuerza.get('activos') or 0}/{fuerza.get('total') or 0} activos",
                ],
            ]
        ),
        Spacer(1, 0.25 * cm),
        _analysis_box(_analisis_kpis(kpis), analysis_style),
    ]
    if resumen:
        header_block.append(Spacer(1, 0.25 * cm))
        header_block.append(Paragraph("Resumen ejecutivo", h2))
        if isinstance(resumen, list):
            for item in resumen:
                texto = item.get("texto") if isinstance(item, dict) else str(item)
                if texto:
                    header_block.append(Paragraph(f"• {texto}", normal))
                    header_block.append(Spacer(1, 0.06 * cm))
        else:
            header_block.append(Paragraph(str(resumen), normal))
    body.append(KeepTogether(header_block))

    # —— Tipología ——
    tip_rows = [["Tipo de delito", "Total", "%"]] + [
        [r.get("tipo_delito") or "—", str(r.get("total") or 0), f"{r.get('pct') or 0}%"]
        for r in tip[:10]
    ]
    if len(tip_rows) == 1:
        tip_rows.append(["Sin datos", "0", "0%"])
    tip_main: list[Any] = [
        Paragraph("Tipología criminal", h2),
        Spacer(1, 0.08 * cm),
    ]
    chart_tip = _tipologia_bar_chart(tip)
    if chart_tip:
        tip_main.append(chart_tip)
        tip_main.append(Spacer(1, 0.18 * cm))
    tip_main.append(_analysis_box(_analisis_tipologia(tip), analysis_style))
    body.append(PageBreak())
    body.append(KeepTogether(tip_main))
    body.append(Spacer(1, 0.2 * cm))
    body.append(_styled_data_table(tip_rows, [11 * cm, 3 * cm, 3 * cm]))

    # —— Evolución ——
    evo_flow: list[Any] = [
        Paragraph("Evolución del delito en el tiempo", h2),
        Spacer(1, 0.08 * cm),
    ]
    chart_evo = _evolucion_line_chart(evo)
    if chart_evo:
        evo_flow.append(chart_evo)
    else:
        evo_flow.append(Paragraph("Sin serie temporal suficiente para graficar.", small))
    evo_flow.append(Spacer(1, 0.18 * cm))
    evo_flow.append(_analysis_box(_analisis_evolucion(evo), analysis_style))
    body.append(PageBreak())
    body.append(KeepTogether(evo_flow))

    # —— Ranking ——
    rank_rows = [["Distrito / sector", "Incidentes", "Nivel"]] + [
        [
            r.get("distrito") or r.get("nombre") or "—",
            str(r.get("total") or r.get("incidentes") or 0),
            str(r.get("nivel") or "—").capitalize(),
        ]
        for r in ranking[:10]
    ]
    if len(rank_rows) == 1:
        rank_rows.append(["Sin datos", "0", "—"])
    rank_main: list[Any] = [
        Paragraph("Ranking de distritos / sectores", h2),
        Spacer(1, 0.08 * cm),
    ]
    chart_rank = _ranking_bar_chart(ranking)
    if chart_rank:
        rank_main.append(chart_rank)
        rank_main.append(Spacer(1, 0.18 * cm))
    rank_main.append(_analysis_box(_analisis_ranking(ranking), analysis_style))
    body.append(PageBreak())
    body.append(KeepTogether(rank_main))
    body.append(Spacer(1, 0.2 * cm))
    body.append(_styled_data_table(rank_rows, [11 * cm, 3 * cm, 3 * cm]))
    body.append(Spacer(1, 0.35 * cm))
    body.append(HRFlowable(width="100%", thickness=0.6, color=LINE, spaceAfter=4))
    body.append(
        Paragraph(
            "Documento generado desde el Dashboard de Inteligencia Táctica · CrimeTrack · Uso interno institucional",
            small,
        )
    )

    doc.build(body)
    return buffer.getvalue()


# ——— PDF pestaña Mapa de Calor ———

HEAT_GREEN = colors.HexColor("#22c55e")
HEAT_ORANGE = colors.HexColor("#f59e0b")
HEAT_RED = colors.HexColor("#b91c1c")
HEAT_RGB = {
    "alta": (185, 28, 28),
    "media": (245, 158, 11),
    "baja": (34, 197, 94),
}

_MAP_UA = "CrimeTrack-PDF/1.0 (police-management; institutional-report)"
_TILE_SIZE = 256


def _heat_color(peso: float, max_peso: float):
    t = min(1.0, peso / max_peso) if max_peso > 0 else 0.3
    if t > 0.7:
        return HEAT_RED
    if t > 0.4:
        return HEAT_ORANGE
    return HEAT_GREEN


def _heat_rgb(peso: float, max_peso: float) -> tuple[int, int, int]:
    t = min(1.0, peso / max_peso) if max_peso > 0 else 0.3
    if t > 0.7:
        return HEAT_RGB["alta"]
    if t > 0.4:
        return HEAT_RGB["media"]
    return HEAT_RGB["baja"]


def _latlon_to_world_px(lat: float, lon: float, zoom: int) -> tuple[float, float]:
    n = 2.0**zoom
    x = (lon + 180.0) / 360.0 * n * _TILE_SIZE
    lat_r = math.radians(lat)
    y = (1.0 - math.asinh(math.tan(lat_r)) / math.pi) / 2.0 * n * _TILE_SIZE
    return x, y


def _choose_zoom(min_lat, max_lat, min_lon, max_lon, target_w=900, target_h=520) -> int:
    for z in range(16, 9, -1):
        x0, y0 = _latlon_to_world_px(max_lat, min_lon, z)
        x1, y1 = _latlon_to_world_px(min_lat, max_lon, z)
        if abs(x1 - x0) <= target_w * 1.35 and abs(y1 - y0) <= target_h * 1.35:
            return z
    return 11


def _fetch_map_tile(z: int, x: int, y: int):
    from PIL import Image

    urls = [
        f"https://a.basemaps.cartocdn.com/rastertiles/voyager/{z}/{x}/{y}.png",
        f"https://b.basemaps.cartocdn.com/rastertiles/voyager/{z}/{x}/{y}.png",
        f"https://tile.openstreetmap.org/{z}/{x}/{y}.png",
    ]
    last_err = None
    for url in urls:
        try:
            req = urllib.request.Request(url, headers={"User-Agent": _MAP_UA})
            with urllib.request.urlopen(req, timeout=6) as resp:
                return Image.open(io.BytesIO(resp.read())).convert("RGB")
        except Exception as exc:  # noqa: BLE001
            last_err = exc
            continue
    raise RuntimeError(f"No se pudo descargar tesela {z}/{x}/{y}: {last_err}")


def _build_basemap_image(
    min_lat: float,
    max_lat: float,
    min_lon: float,
    max_lon: float,
    *,
    out_w: int = 980,
    out_h: int = 520,
):
    """Compone un mapa real (calles/nombres) con teselas Carto/OSM."""
    zoom = _choose_zoom(min_lat, max_lat, min_lon, max_lon, out_w, out_h)
    # Limita teselas para no saturar la red
    for _ in range(4):
        wx0, wy0 = _latlon_to_world_px(max_lat, min_lon, zoom)
        wx1, wy1 = _latlon_to_world_px(min_lat, max_lon, zoom)
        pad = 48
        cols = int(math.floor((wx1 + pad) / _TILE_SIZE) - math.floor((wx0 - pad) / _TILE_SIZE) + 1)
        rows = int(math.floor((wy1 + pad) / _TILE_SIZE) - math.floor((wy0 - pad) / _TILE_SIZE) + 1)
        if cols * rows <= 40:
            break
        zoom = max(10, zoom - 1)
    return _build_basemap_image_at_zoom(
        min_lat, max_lat, min_lon, max_lon, zoom, out_w, out_h
    )


def _build_basemap_image_at_zoom(
    min_lat, max_lat, min_lon, max_lon, zoom, out_w, out_h
):
    from PIL import Image

    wx0, wy0 = _latlon_to_world_px(max_lat, min_lon, zoom)
    wx1, wy1 = _latlon_to_world_px(min_lat, max_lon, zoom)
    pad = 48
    wx0 -= pad
    wy0 -= pad
    wx1 += pad
    wy1 += pad

    tx0 = int(math.floor(wx0 / _TILE_SIZE))
    ty0 = int(math.floor(wy0 / _TILE_SIZE))
    tx1 = int(math.floor(wx1 / _TILE_SIZE))
    ty1 = int(math.floor(wy1 / _TILE_SIZE))
    n_tiles = 2**zoom
    tx0 = max(0, min(tx0, n_tiles - 1))
    ty0 = max(0, min(ty0, n_tiles - 1))
    tx1 = max(0, min(tx1, n_tiles - 1))
    ty1 = max(0, min(ty1, n_tiles - 1))
    cols = tx1 - tx0 + 1
    rows = ty1 - ty0 + 1

    mosaic = Image.new("RGB", (cols * _TILE_SIZE, rows * _TILE_SIZE), (241, 245, 249))
    jobs = [(x, y) for y in range(ty0, ty1 + 1) for x in range(tx0, tx1 + 1)]

    def _one(xy):
        x, y = xy
        try:
            return x, y, _fetch_map_tile(zoom, x, y)
        except Exception:
            blank = Image.new("RGB", (_TILE_SIZE, _TILE_SIZE), (226, 232, 240))
            return x, y, blank

    with ThreadPoolExecutor(max_workers=8) as pool:
        for fut in as_completed([pool.submit(_one, j) for j in jobs]):
            x, y, tile = fut.result()
            mosaic.paste(tile, ((x - tx0) * _TILE_SIZE, (y - ty0) * _TILE_SIZE))

    # recorte al bbox
    left = int(wx0 - tx0 * _TILE_SIZE)
    top = int(wy0 - ty0 * _TILE_SIZE)
    right = int(wx1 - tx0 * _TILE_SIZE)
    bottom = int(wy1 - ty0 * _TILE_SIZE)
    left = max(0, left)
    top = max(0, top)
    right = min(mosaic.width, max(left + 10, right))
    bottom = min(mosaic.height, max(top + 10, bottom))
    cropped = mosaic.crop((left, top, right, bottom))
    cropped = cropped.resize((out_w, out_h), Image.Resampling.LANCZOS)
    return cropped, zoom, (wx0, wy0, wx1, wy1)


def _latlon_to_image_xy(lat, lon, zoom, world_bbox, img_w, img_h):
    wx0, wy0, wx1, wy1 = world_bbox
    wx, wy = _latlon_to_world_px(lat, lon, zoom)
    x = (wx - wx0) / max(wx1 - wx0, 1e-6) * img_w
    y = (wy - wy0) / max(wy1 - wy0, 1e-6) * img_h
    return x, y


def _render_mapa_calor_png(puntos: list[dict]) -> bytes | None:
    """Mapa real (calles/nombres) + focos de calor etiquetados."""
    if not puntos:
        return None
    try:
        from PIL import Image, ImageDraw, ImageFont
    except ImportError:
        return None

    sample = sorted(puntos, key=lambda p: int(p.get("peso") or 0), reverse=True)[:200]
    lats = [float(p["latitud"]) for p in sample]
    lons = [float(p["longitud"]) for p in sample]
    min_lat, max_lat = min(lats), max(lats)
    min_lon, max_lon = min(lons), max(lons)
    if abs(max_lat - min_lat) < 0.004:
        min_lat -= 0.01
        max_lat += 0.01
    if abs(max_lon - min_lon) < 0.004:
        min_lon -= 0.01
        max_lon += 0.01

    out_w, out_h = 980, 520
    try:
        base, zoom, world_bbox = _build_basemap_image(
            min_lat, max_lat, min_lon, max_lon, out_w=out_w, out_h=out_h
        )
    except Exception:
        # Fallback: fondo neutro si no hay red a teselas
        base = Image.new("RGB", (out_w, out_h), (241, 245, 249))
        zoom = 12
        world_bbox = (
            *_latlon_to_world_px(max_lat, min_lon, zoom),
            *_latlon_to_world_px(min_lat, max_lon, zoom),
        )

    overlay = Image.new("RGBA", base.size, (0, 0, 0, 0))
    draw = ImageDraw.Draw(overlay, "RGBA")
    try:
        font = ImageFont.truetype("DejaVuSans.ttf", 13)
        font_sm = ImageFont.truetype("DejaVuSans.ttf", 11)
    except Exception:
        try:
            font = ImageFont.truetype("arial.ttf", 13)
            font_sm = ImageFont.truetype("arial.ttf", 11)
        except Exception:
            font = ImageFont.load_default()
            font_sm = font

    max_peso = max(int(p.get("peso") or 0) for p in sample) or 1

    # focos (primero los más débiles)
    for p in reversed(sample):
        lat = float(p["latitud"])
        lon = float(p["longitud"])
        peso = int(p.get("peso") or 0)
        x, y = _latlon_to_image_xy(lat, lon, zoom, world_bbox, out_w, out_h)
        r = 7 + min(18, (peso / max_peso) * 18)
        rgb = _heat_rgb(peso, max_peso)
        draw.ellipse(
            (x - r, y - r, x + r, y + r),
            fill=(*rgb, 110),
            outline=(*rgb, 220),
            width=2,
        )

    # etiquetas de los focos más peligrosos (nombres / tipología)
    top_labels = sample[:8]
    for i, p in enumerate(top_labels, start=1):
        lat = float(p["latitud"])
        lon = float(p["longitud"])
        x, y = _latlon_to_image_xy(lat, lon, zoom, world_bbox, out_w, out_h)
        label = f"{i}. {(p.get('tipo_delito') or 'Foco')[:22]}"
        # globo blanco legible
        tb = draw.textbbox((0, 0), label, font=font_sm)
        tw, th = tb[2] - tb[0], tb[3] - tb[1]
        lx = min(max(6, x + 10), out_w - tw - 10)
        ly = min(max(6, y - th - 12), out_h - th - 8)
        draw.rounded_rectangle(
            (lx - 4, ly - 3, lx + tw + 4, ly + th + 3),
            radius=4,
            fill=(255, 255, 255, 230),
            outline=(91, 61, 154, 220),
            width=1,
        )
        draw.text((lx, ly), label, fill=(51, 65, 85, 255), font=font_sm)
        # marcador número
        draw.ellipse((x - 8, y - 8, x + 8, y + 8), fill=(91, 61, 154, 240), outline=(255, 255, 255, 255))
        num = str(i)
        nb = draw.textbbox((0, 0), num, font=font_sm)
        nw, nh = nb[2] - nb[0], nb[3] - nb[1]
        draw.text((x - nw / 2, y - nh / 2 - 1), num, fill=(255, 255, 255, 255), font=font_sm)

    # leyenda
    legend_items = [
        ("Baja", HEAT_RGB["baja"]),
        ("Media", HEAT_RGB["media"]),
        ("Alta densidad", HEAT_RGB["alta"]),
    ]
    lx0, ly0 = 14, out_h - 34
    draw.rounded_rectangle(
        (8, out_h - 42, 290, out_h - 8),
        radius=6,
        fill=(255, 255, 255, 220),
        outline=(226, 232, 240, 255),
    )
    cx = lx0
    for lab, rgb in legend_items:
        draw.ellipse((cx, ly0, cx + 12, ly0 + 12), fill=(*rgb, 200), outline=(*rgb, 255))
        draw.text((cx + 16, ly0 - 1), lab, fill=(51, 65, 85, 255), font=font_sm)
        cx += 90

    composed = Image.alpha_composite(base.convert("RGBA"), overlay).convert("RGB")
    buf = io.BytesIO()
    composed.save(buf, format="PNG", optimize=True)
    return buf.getvalue()


def _mapa_flowable(puntos: list[dict]):
    """Imagen de mapa real o, si falla la red, dibujo vectorial de respaldo."""
    png = _render_mapa_calor_png(puntos)
    if png:
        img = RLImage(io.BytesIO(png), width=26 * cm, height=12.2 * cm)
        img.hAlign = "CENTER"
        return img
    return _mapa_scatter_drawing(puntos)


def _mapa_scatter_drawing(puntos: list[dict]) -> Drawing | None:
    if not puntos:
        return None
    sample = sorted(puntos, key=lambda p: int(p.get("peso") or 0), reverse=True)[:180]
    lats = [float(p["latitud"]) for p in sample]
    lons = [float(p["longitud"]) for p in sample]
    min_lat, max_lat = min(lats), max(lats)
    min_lon, max_lon = min(lons), max(lons)
    pad_lat = max((max_lat - min_lat) * 0.08, 0.002)
    pad_lon = max((max_lon - min_lon) * 0.08, 0.002)
    min_lat -= pad_lat
    max_lat += pad_lat
    min_lon -= pad_lon
    max_lon += pad_lon
    span_lat = max(max_lat - min_lat, 1e-6)
    span_lon = max(max_lon - min_lon, 1e-6)
    max_peso = max(int(p.get("peso") or 0) for p in sample) or 1

    w, h, margin = 520, 250, 18
    drawing = Drawing(w, h)
    drawing.add(
        Rect(0, 0, w, h, fillColor=colors.HexColor("#f1f5f9"), strokeColor=LINE, strokeWidth=0.8)
    )
    for p in reversed(sample):
        lat = float(p["latitud"])
        lon = float(p["longitud"])
        peso = int(p.get("peso") or 0)
        x = margin + ((lon - min_lon) / span_lon) * (w - 2 * margin)
        y = margin + ((lat - min_lat) / span_lat) * (h - 2 * margin)
        r = 3.5 + min(10.0, (peso / max_peso) * 10.0)
        col = _heat_color(peso, max_peso)
        drawing.add(Circle(x, y, r, fillColor=col, strokeColor=col, strokeWidth=0.6))
    return drawing


def _radar_drawing(dias: list[dict]) -> Drawing | None:
    data = dias or []
    if not data:
        return None
    w = h = 260
    cx = cy = w / 2
    r = 85
    max_v = max([int(d.get("total") or 0) for d in data] + [1])
    n = len(data)
    drawing = Drawing(w, h)
    drawing.add(Circle(cx, cy, r, fillColor=None, strokeColor=LINE, strokeWidth=0.8))
    drawing.add(Circle(cx, cy, r * 0.66, fillColor=None, strokeColor=LINE, strokeWidth=0.5))
    drawing.add(Circle(cx, cy, r * 0.33, fillColor=None, strokeColor=LINE, strokeWidth=0.5))

    pts = []
    for i, d in enumerate(data):
        ang = (-math.pi / 2) + (i / n) * math.pi * 2
        total = int(d.get("total") or 0)
        rr = (total / max_v) * r
        x = cx + math.cos(ang) * rr
        y = cy + math.sin(ang) * rr
        ax = cx + math.cos(ang) * r
        ay = cy + math.sin(ang) * r
        lx = cx + math.cos(ang) * (r + 18)
        ly = cy + math.sin(ang) * (r + 18)
        pts.append((x, y))
        drawing.add(Line(cx, cy, ax, ay, strokeColor=colors.HexColor("#eef2f7"), strokeWidth=0.8))
        label = str(d.get("label") or "")[:3]
        drawing.add(String(lx - 8, ly - 3, label, fontSize=8, fillColor=SLATE))
        drawing.add(Circle(x, y, 4, fillColor=PURPLE, strokeColor=PURPLE_DARK, strokeWidth=0.5))

    if len(pts) >= 3:
        flat = []
        for x, y in pts:
            flat.extend([x, y])
        drawing.add(
            Polygon(
                flat,
                fillColor=colors.Color(124 / 255, 92 / 255, 191 / 255, alpha=0.22),
                strokeColor=PURPLE,
                strokeWidth=2,
            )
        )
    return drawing


def _top_focos(puntos: list[dict], limit: int = 12) -> list[dict]:
    ranked = sorted(puntos or [], key=lambda p: int(p.get("peso") or 0), reverse=True)
    return ranked[:limit]


def _densidad_label(peso: float, max_peso: float) -> str:
    col = _heat_color(peso, max_peso)
    if col == HEAT_RED:
        return "Alta"
    if col == HEAT_ORANGE:
        return "Media"
    return "Baja"


def _table_section(title: str, table: Table, h2_style: ParagraphStyle) -> KeepTogether:
    """Título + tabla inseparables (no parten entre hojas)."""
    return KeepTogether(
        [
            Paragraph(title, h2_style),
            Spacer(1, 0.12 * cm),
            table,
        ]
    )


def _analisis_mapa(puntos: list[dict], total_puntos: int) -> str:
    if not puntos:
        return (
            "No hay focos georreferenciados con los filtros actuales. "
            "Verifique el periodo o la tipología antes de planificar operativos territoriales."
        )
    top = _top_focos(puntos, 3)
    max_peso = max(int(p.get("peso") or 0) for p in puntos) or 1
    altos = sum(1 for p in puntos if (int(p.get("peso") or 0) / max_peso) > 0.7)
    tipos: dict[str, int] = {}
    for p in puntos:
        t = p.get("tipo_delito") or "Sin clasificar"
        tipos[t] = tipos.get(t, 0) + int(p.get("peso") or 0)
    top_tipo = max(tipos.items(), key=lambda x: x[1])[0] if tipos else "—"
    t0 = top[0]
    return (
        f"El mapa muestra <b>{total_puntos}</b> focos sobre cartografía real (calles y barrios). "
        f"Hay <b>{altos}</b> punto(s) de alta densidad (rojo). "
        f"El foco #1 es <b>{t0.get('tipo_delito') or 'delito'}</b> "
        f"(peso {t0.get('peso')}, lat {float(t0.get('latitud')):.4f}, lon {float(t0.get('longitud')):.4f}). "
        f"Tipología dominante: <b>{top_tipo}</b>. "
        "Use los números del mapa para priorizar patrullaje en esas ubicaciones concretas."
    )


def _analisis_radar(radar: dict) -> str:
    dias = radar.get("dias") or []
    picos = radar.get("picos") or []
    if not dias:
        return (
            "Sin datos del reloj criminológico. No es posible programar refuerzos por día/hora "
            "hasta disponer de serie temporal."
        )
    top_dia = max(dias, key=lambda d: int(d.get("total") or 0))
    texto = (
        f"El radar resume la carga por día de la semana. El día de mayor incidencia es "
        f"<b>{top_dia.get('label')}</b> con <b>{top_dia.get('total')}</b> incidentes: "
        "programe allí mayor disponibilidad de personal."
    )
    if picos:
        p = picos[0]
        texto += (
            f" El pico más agudo es <b>{p.get('dia_label')} {int(p.get('hora') or 0):02d}:00</b> "
            f"({p.get('total')} incidentes): ideal para operativos focalizados en esa franja."
        )
        if len(picos) > 1:
            extras = ", ".join(
                f"{x.get('dia_label')} {int(x.get('hora') or 0):02d}:00"
                for x in picos[1:3]
            )
            texto += f" Otros picos relevantes: {extras}."
    texto += " Use el reloj para decidir turnos y el mapa para decidir dónde desplegarlos."
    return texto


def build_mapa_calor_pdf(
    *,
    emisor: str,
    jurisdiccion: dict[str, Any],
    filtros: dict[str, Any],
    mapa: dict[str, Any],
    radar: dict[str, Any] | None = None,
) -> bytes:
    """PDF de la pestaña Mapa de Calor: focos peligrosos + reloj criminológico + análisis."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=1.4 * cm,
        rightMargin=1.4 * cm,
        topMargin=1.1 * cm,
        bottomMargin=1.1 * cm,
    )
    styles = getSampleStyleSheet()
    title = ParagraphStyle(
        "McTitle",
        parent=styles["Heading1"],
        fontSize=17,
        textColor=PURPLE_DARK,
        spaceAfter=2,
        leading=20,
    )
    kicker = ParagraphStyle(
        "McKicker",
        parent=styles["Normal"],
        fontSize=9,
        textColor=PURPLE,
        fontName="Helvetica-Bold",
        spaceAfter=2,
    )
    h2 = ParagraphStyle(
        "McH2",
        parent=styles["Heading2"],
        fontSize=12,
        textColor=SLATE,
        spaceBefore=0,
        spaceAfter=6,
    )
    normal = ParagraphStyle(
        "McNormal",
        parent=styles["Normal"],
        fontSize=9,
        textColor=SLATE,
        leading=12,
    )
    analysis_style = ParagraphStyle(
        "McAnalysis",
        parent=styles["Normal"],
        fontSize=8.5,
        textColor=SLATE,
        leading=11.5,
    )
    small = ParagraphStyle(
        "McSmall",
        parent=styles["Normal"],
        fontSize=8,
        textColor=MUTED,
        leading=10,
    )

    radar = radar or {}
    puntos = mapa.get("puntos") or []
    total_puntos = int(mapa.get("total_puntos") or len(puntos))
    generado = datetime.now().strftime("%Y-%m-%d %H:%M")
    body: list[Any] = []

    # —— Hoja 1: encabezado + mapa + análisis ——
    map_visual = _mapa_flowable(puntos)
    body.extend(
        [
            Paragraph("INTELIGENCIA TÁCTICA · MAPA DE CALOR", kicker),
            Paragraph(
                f"Zonas de riesgo — {jurisdiccion.get('nombre') or 'Jurisdicción'}",
                title,
            ),
            Paragraph(
                f"<b>Periodo:</b> {filtros.get('fecha_desde') or '—'} → {filtros.get('fecha_hasta') or '—'} "
                f"&nbsp;&nbsp;|&nbsp;&nbsp; <b>Distrito:</b> {filtros.get('distrito') or 'Todos'} "
                f"&nbsp;&nbsp;|&nbsp;&nbsp; <b>Delito:</b> {filtros.get('tipo_delito') or 'Todos'}",
                normal,
            ),
            Paragraph(
                f"Elaborado por <b>{emisor}</b> · Generado {generado} · "
                f"{total_puntos} focos georreferenciados · Código {jurisdiccion.get('codigo') or '—'}",
                small,
            ),
            Spacer(1, 0.12 * cm),
            HRFlowable(width="100%", thickness=1.2, color=PURPLE, spaceAfter=6),
            Paragraph(f"Mapa de calor · {total_puntos} focos", h2),
            Spacer(1, 0.06 * cm),
        ]
    )
    if map_visual:
        body.append(map_visual)
    else:
        body.append(Paragraph("Sin coordenadas para graficar focos.", small))
    body.append(Spacer(1, 0.15 * cm))
    body.append(_analysis_box(_analisis_mapa(puntos, total_puntos), analysis_style))

    # —— Hoja 2: tabla de focos (completa, sin partir) ——
    focos = _top_focos(puntos, 12)
    max_peso = max((int(x.get("peso") or 0) for x in focos), default=1) or 1
    foco_rows = [["#", "Tipo de delito", "Peso", "Latitud", "Longitud", "Densidad"]] + [
        [
            str(i + 1),
            (f.get("tipo_delito") or "—")[:28],
            str(f.get("peso") or 0),
            f"{float(f.get('latitud') or 0):.5f}",
            f"{float(f.get('longitud') or 0):.5f}",
            _densidad_label(float(f.get("peso") or 0), max_peso),
        ]
        for i, f in enumerate(focos)
    ]
    if len(foco_rows) == 1:
        foco_rows.append(["—", "Sin focos", "0", "—", "—", "—"])
    body.append(PageBreak())
    body.append(
        _table_section(
            "Zonas peligrosas prioritarias (top focos)",
            _styled_data_table(
                foco_rows, [1.2 * cm, 7 * cm, 2 * cm, 3.5 * cm, 3.5 * cm, 2.5 * cm]
            ),
            h2,
        )
    )

    # —— Hoja 3: reloj + análisis ——
    radar_block: list[Any] = [
        Paragraph("Reloj criminológico", h2),
        Spacer(1, 0.08 * cm),
        Paragraph(
            "Distribución de incidentes por día de la semana. Los picos indican franjas críticas.",
            small,
        ),
        Spacer(1, 0.1 * cm),
    ]
    radar_draw = _radar_drawing(radar.get("dias") or [])
    if radar_draw:
        radar_block.append(radar_draw)
    else:
        radar_block.append(Paragraph("Sin datos de radar.", small))
    radar_block.append(Spacer(1, 0.18 * cm))
    radar_block.append(_analysis_box(_analisis_radar(radar), analysis_style))
    body.append(PageBreak())
    body.append(KeepTogether(radar_block))

    # —— Hoja 4: tabla picos (sola) ——
    picos = radar.get("picos") or []
    pico_rows = [["Día", "Hora", "Franja", "Incidentes"]] + [
        [
            p.get("dia_label") or "—",
            f"{int(p.get('hora') or 0):02d}:00",
            p.get("franja_label") or "—",
            str(p.get("total") or 0),
        ]
        for p in picos
    ]
    if len(pico_rows) == 1:
        pico_rows.append(["Sin picos", "—", "—", "0"])
    body.append(PageBreak())
    body.append(
        _table_section(
            "Picos detectados (día + hora)",
            _styled_data_table(pico_rows, [4 * cm, 3 * cm, 7 * cm, 3 * cm]),
            h2,
        )
    )

    # —— Hoja 5: carga por día (sola) ——
    dias = radar.get("dias") or []
    if dias:
        dia_rows = [["Día", "Total incidentes"]] + [
            [d.get("label") or "—", str(d.get("total") or 0)] for d in dias
        ]
        body.append(PageBreak())
        body.append(
            KeepTogether(
                [
                    _table_section(
                        "Carga por día de la semana",
                        _styled_data_table(dia_rows, [8 * cm, 5 * cm]),
                        h2,
                    ),
                    Spacer(1, 0.4 * cm),
                    HRFlowable(width="100%", thickness=0.6, color=LINE, spaceAfter=4),
                    Paragraph(
                        "Documento generado desde la pestaña Mapa de Calor · CrimeTrack · Uso interno institucional",
                        small,
                    ),
                ]
            )
        )
    else:
        body.append(Spacer(1, 0.35 * cm))
        body.append(HRFlowable(width="100%", thickness=0.6, color=LINE, spaceAfter=4))
        body.append(
            Paragraph(
                "Documento generado desde la pestaña Mapa de Calor · CrimeTrack · Uso interno institucional",
                small,
            )
        )

    doc.build(body)
    return buffer.getvalue()


# ——— PDF pestaña Ranking Distritos ———

CUAD_COLORS = {
    "rojo": colors.HexColor("#ef4444"),
    "verde": colors.HexColor("#22c55e"),
    "amarillo": colors.HexColor("#f59e0b"),
    "neutro": colors.HexColor("#7c5cbf"),
}

CUAD_LABELS = {
    "rojo": "Muchos delitos / pocos arrestos",
    "verde": "Alta efectividad",
    "amarillo": "Equilibrado",
    "neutro": "Neutro",
}


def _scatter_eficiencia_drawing(ranking: list[dict]) -> Drawing | None:
    rows = ranking or []
    if not rows:
        return None

    def _short(name: str) -> str:
        raw = (name or "").strip()
        if not raw:
            return "—"
        first = raw.split("—")[0].split("–")[0].split(" - ")[0].strip()
        return first[:22] + ("…" if len(first) > 22 else "")

    def _domain(values: list[float]) -> tuple[float, float]:
        if not values:
            return 0.0, 1.0
        min_v = float(min(values))
        max_v = float(max(values))
        if max_v == min_v:
            pad = max(abs(max_v) * 0.35, 2.0)
            min_v -= pad
            max_v += pad
        else:
            span = max_v - min_v
            pad = max(span * 0.35, 1.0)
            min_v -= pad
            max_v += pad
        min_v = max(0.0, min_v)
        if max_v <= min_v:
            max_v = min_v + 1.0
        return min_v, max_v

    w, h = 560, 300
    pad_l, pad_r, pad_t, pad_b = 52, 110, 28, 46
    inner_w = w - pad_l - pad_r
    inner_h = h - pad_t - pad_b
    xs = [float(r.get("delitos") or 0) for r in rows]
    ys = [float(r.get("arrestos") or 0) for r in rows]
    min_x, max_x = _domain(xs)
    min_y, max_y = _domain(ys)
    span_x = max(max_x - min_x, 1e-6)
    span_y = max(max_y - min_y, 1e-6)

    def to_x(v: float) -> float:
        return pad_l + ((v - min_x) / span_x) * inner_w

    def to_y(v: float) -> float:
        return pad_t + inner_h - ((v - min_y) / span_y) * inner_h

    mid_x = sorted(xs)[len(xs) // 2]
    mid_y = sorted(ys)[len(ys) // 2]
    x_ticks = [min_x, (min_x + max_x) / 2, max_x]
    y_ticks = [min_y, (min_y + max_y) / 2, max_y]

    drawing = Drawing(w, h)
    drawing.add(
        Rect(0, 0, w, h, fillColor=colors.HexColor("#fafbfc"), strokeColor=LINE, strokeWidth=0.8)
    )
    for tick in x_ticks:
        drawing.add(
            Line(
                to_x(tick),
                pad_t,
                to_x(tick),
                pad_t + inner_h,
                strokeColor=colors.HexColor("#eef2f7"),
                strokeWidth=0.6,
            )
        )
    for tick in y_ticks:
        drawing.add(
            Line(
                pad_l,
                to_y(tick),
                pad_l + inner_w,
                to_y(tick),
                strokeColor=colors.HexColor("#eef2f7"),
                strokeWidth=0.6,
            )
        )
    drawing.add(
        Line(
            pad_l,
            pad_t + inner_h,
            pad_l + inner_w,
            pad_t + inner_h,
            strokeColor=colors.HexColor("#94a3b8"),
            strokeWidth=1.2,
        )
    )
    drawing.add(
        Line(
            pad_l,
            pad_t,
            pad_l,
            pad_t + inner_h,
            strokeColor=colors.HexColor("#94a3b8"),
            strokeWidth=1.2,
        )
    )
    drawing.add(
        Line(
            to_x(mid_x),
            pad_t,
            to_x(mid_x),
            pad_t + inner_h,
            strokeColor=colors.HexColor("#cbd5e1"),
            strokeWidth=0.8,
        )
    )
    drawing.add(
        Line(
            pad_l,
            to_y(mid_y),
            pad_l + inner_w,
            to_y(mid_y),
            strokeColor=colors.HexColor("#cbd5e1"),
            strokeWidth=0.8,
        )
    )

    for tick in x_ticks:
        drawing.add(
            String(
                to_x(tick) - 8,
                h - 28,
                str(int(round(tick))),
                fontSize=8,
                fillColor=MUTED,
            )
        )
    for tick in y_ticks:
        drawing.add(
            String(
                pad_l - 28,
                to_y(tick) - 2,
                str(int(round(tick))),
                fontSize=8,
                fillColor=MUTED,
            )
        )

    drawing.add(
        String(pad_l + inner_w / 2 - 55, 8, "Delitos reportados →", fontSize=9, fillColor=SLATE)
    )
    drawing.add(String(6, h / 2 + 20, "Arrestos", fontSize=8, fillColor=SLATE))
    drawing.add(String(6, h / 2 + 8, "(efectividad)", fontSize=7, fillColor=SLATE))

    # puntos + etiquetas (con anti-solape vertical)
    placed = []
    for i, r in enumerate(rows):
        delitos = float(r.get("delitos") or 0)
        arrestos = float(r.get("arrestos") or 0)
        x = to_x(delitos)
        y = to_y(arrestos)
        placed.append(
            {
                "i": i + 1,
                "r": r,
                "x": x,
                "y": y,
                "label_left": x > pad_l + inner_w * 0.62,
                "label_y": y,
            }
        )
    placed.sort(key=lambda p: p["label_y"])
    for i in range(1, len(placed)):
        if placed[i]["label_y"] - placed[i - 1]["label_y"] < 16:
            placed[i]["label_y"] = placed[i - 1]["label_y"] + 16

    for p in placed:
        r = p["r"]
        col = CUAD_COLORS.get(r.get("cuadrante") or "neutro", CUAD_COLORS["neutro"])
        drawing.add(Circle(p["x"], p["y"], 8, fillColor=col, strokeColor=colors.white, strokeWidth=1.2))
        drawing.add(
            String(p["x"] - 3, p["y"] - 3, str(p["i"]), fontSize=7, fillColor=colors.white)
        )
        nombre = _short(r.get("distrito") or "—")
        meta = f"{int(r.get('delitos') or 0)} del. · {int(r.get('arrestos') or 0)} arr."
        if p["label_left"]:
            drawing.add(
                String(max(4, p["x"] - 100), p["label_y"] - 2, nombre, fontSize=8, fillColor=SLATE)
            )
            drawing.add(
                String(max(4, p["x"] - 100), p["label_y"] - 12, meta, fontSize=6.5, fillColor=MUTED)
            )
        else:
            drawing.add(String(p["x"] + 12, p["label_y"] - 2, nombre, fontSize=8, fillColor=SLATE))
            drawing.add(String(p["x"] + 12, p["label_y"] - 12, meta, fontSize=6.5, fillColor=MUTED))

    legend = [("rojo", "Crítico"), ("verde", "Efectivo"), ("amarillo", "Equilibrado")]
    lx = w - 210
    for i, (key, lab) in enumerate(legend):
        xx = lx + i * 70
        drawing.add(Circle(xx, h - 10, 4, fillColor=CUAD_COLORS[key], strokeColor=CUAD_COLORS[key]))
        drawing.add(String(xx + 7, h - 13, lab, fontSize=7, fillColor=SLATE))
    return drawing


def _analisis_matriz(ranking: list[dict]) -> str:
    if not ranking:
        return (
            "Sin distritos en la matriz. Ajuste el periodo o filtros para evaluar "
            "la relación delitos × arrestos."
        )
    rojos = [r for r in ranking if r.get("cuadrante") == "rojo"]
    verdes = [r for r in ranking if r.get("cuadrante") == "verde"]
    amarillos = [r for r in ranking if r.get("cuadrante") == "amarillo"]
    top = ranking[0]
    texto = (
        f"La matriz ubica cada distrito según <b>delitos reportados</b> (eje X) y "
        f"<b>arrestos</b> (eje Y). El líder en volumen es <b>{top.get('distrito')}</b> "
        f"({top.get('delitos')} delitos / {top.get('arrestos')} arrestos)."
    )
    if rojos:
        names = ", ".join(r.get("distrito") or "—" for r in rojos[:3])
        texto += (
            f" Cuadrante crítico (rojo): <b>{names}</b> — mucha incidencia y baja captura; "
            "refuerce investigación y operativos focalizados."
        )
    if verdes:
        names = ", ".join(r.get("distrito") or "—" for r in verdes[:3])
        texto += f" Alta efectividad (verde): <b>{names}</b> — modele sus prácticas en otros sectores."
    if amarillos:
        texto += (
            f" {len(amarillos)} distrito(s) equilibrado(s): mantenga el ritmo y vigile desviaciones."
        )
    texto += (
        " Interprete puntos a la derecha como mayor carga delictiva; puntos arriba, mejor capacidad de respuesta."
    )
    return texto


def _analisis_leaderboard(ranking: list[dict]) -> str:
    if not ranking:
        return "No hay leaderboard disponible con los filtros actuales."
    top3 = ranking[:3]
    lineas = []
    for i, r in enumerate(top3, 1):
        tend = int(r.get("tendencia") or 0)
        tend_txt = "al alza" if tend > 0 else ("a la baja" if tend < 0 else "estable")
        lineas.append(
            f"#{i} <b>{r.get('distrito')}</b> ({r.get('delitos')} delitos, "
            f"{r.get('arrestos')} arrestos, tendencia 7 días {tend_txt})"
        )
    texto = "Priorice supervisión en el Top 3: " + "; ".join(lineas) + ". "
    peores = [r for r in ranking if r.get("cuadrante") == "rojo"]
    if peores:
        texto += (
            f"Atención especial a <b>{peores[0].get('distrito')}</b> por su posición crítica "
            "en la matriz. "
        )
    mejores = [r for r in ranking if r.get("cuadrante") == "verde"]
    if mejores:
        texto += (
            f"Reconozca el desempeño de <b>{mejores[0].get('distrito')}</b> (alta efectividad). "
        )
    texto += (
        "Use el ranking para redistribuir escuadras: más apoyo a los de mayor volumen "
        "y menor ratio de arrestos."
    )
    return texto


def build_ranking_distritos_pdf(
    *,
    emisor: str,
    jurisdiccion: dict[str, Any],
    filtros: dict[str, Any],
    ranking: list[dict] | None = None,
) -> bytes:
    """PDF de Ranking Distritos: matriz de eficiencia + leaderboard + análisis."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=1.4 * cm,
        rightMargin=1.4 * cm,
        topMargin=1.1 * cm,
        bottomMargin=1.1 * cm,
    )
    styles = getSampleStyleSheet()
    title = ParagraphStyle(
        "RkTitle",
        parent=styles["Heading1"],
        fontSize=17,
        textColor=PURPLE_DARK,
        spaceAfter=2,
        leading=20,
    )
    kicker = ParagraphStyle(
        "RkKicker",
        parent=styles["Normal"],
        fontSize=9,
        textColor=PURPLE,
        fontName="Helvetica-Bold",
        spaceAfter=2,
    )
    h2 = ParagraphStyle(
        "RkH2",
        parent=styles["Heading2"],
        fontSize=12,
        textColor=SLATE,
        spaceBefore=0,
        spaceAfter=6,
    )
    normal = ParagraphStyle(
        "RkNormal",
        parent=styles["Normal"],
        fontSize=9,
        textColor=SLATE,
        leading=12,
    )
    analysis_style = ParagraphStyle(
        "RkAnalysis",
        parent=styles["Normal"],
        fontSize=8.5,
        textColor=SLATE,
        leading=11.5,
    )
    small = ParagraphStyle(
        "RkSmall",
        parent=styles["Normal"],
        fontSize=8,
        textColor=MUTED,
        leading=10,
    )

    ranking = ranking or []
    generado = datetime.now().strftime("%Y-%m-%d %H:%M")
    body: list[Any] = []

    body.extend(
        [
            Paragraph("INTELIGENCIA TÁCTICA · RANKING DISTRITOS", kicker),
            Paragraph(
                f"Eficiencia territorial — {jurisdiccion.get('nombre') or 'Jurisdicción'}",
                title,
            ),
            Paragraph(
                f"<b>Periodo:</b> {filtros.get('fecha_desde') or '—'} → {filtros.get('fecha_hasta') or '—'} "
                f"&nbsp;&nbsp;|&nbsp;&nbsp; <b>Distrito:</b> {filtros.get('distrito') or 'Todos'} "
                f"&nbsp;&nbsp;|&nbsp;&nbsp; <b>Delito:</b> {filtros.get('tipo_delito') or 'Todos'}",
                normal,
            ),
            Paragraph(
                f"Elaborado por <b>{emisor}</b> · Generado {generado} · "
                f"{len(ranking)} distrito(s) · Código {jurisdiccion.get('codigo') or '—'}",
                small,
            ),
            Spacer(1, 0.12 * cm),
            HRFlowable(width="100%", thickness=1.2, color=PURPLE, spaceAfter=8),
        ]
    )

    matriz_block: list[Any] = [
        Paragraph("Matriz de Eficiencia (Delitos × Arrestos)", h2),
        Spacer(1, 0.06 * cm),
        Paragraph(
            "Rojo = crítica · Verde = alta efectividad · Amarillo = equilibrado",
            small,
        ),
        Spacer(1, 0.1 * cm),
    ]
    scatter = _scatter_eficiencia_drawing(ranking)
    if scatter:
        matriz_block.append(scatter)
    else:
        matriz_block.append(Paragraph("Sin distritos para graficar.", small))
    matriz_block.append(Spacer(1, 0.18 * cm))
    matriz_block.append(_analysis_box(_analisis_matriz(ranking), analysis_style))
    body.append(KeepTogether(matriz_block))

    board_rows = [["#", "Distrito", "Delitos", "Arrestos", "Nivel", "Cuadrante", "Tend. 7d"]]
    for i, r in enumerate(ranking[:15], 1):
        tend = int(r.get("tendencia") or 0)
        board_rows.append(
            [
                str(i),
                (r.get("distrito") or "—")[:36],
                str(r.get("delitos") or 0),
                str(r.get("arrestos") or 0),
                str(r.get("nivel") or "—").capitalize(),
                CUAD_LABELS.get(r.get("cuadrante") or "neutro", "—")[:22],
                f"{'+' if tend > 0 else ''}{tend}",
            ]
        )
    if len(board_rows) == 1:
        board_rows.append(["—", "Sin datos", "0", "0", "—", "—", "0"])

    body.append(PageBreak())
    leader_block: list[Any] = [
        Paragraph("Leaderboard · Top distritos", h2),
        Spacer(1, 0.08 * cm),
        Paragraph("Ordenado por volumen de delitos. La tendencia refleja los últimos 7 días.", small),
        Spacer(1, 0.12 * cm),
        _styled_data_table(
            board_rows,
            [1.2 * cm, 7.5 * cm, 2.2 * cm, 2.2 * cm, 2.2 * cm, 5.2 * cm, 2 * cm],
        ),
        Spacer(1, 0.2 * cm),
        _analysis_box(_analisis_leaderboard(ranking), analysis_style),
    ]
    body.append(KeepTogether(leader_block))

    if ranking:
        body.append(PageBreak())
        spark_rows_flow: list[Any] = [
            Paragraph("Tendencia reciente por distrito (últimos 7 días)", h2),
            Spacer(1, 0.15 * cm),
        ]
        spark_table_rows = [["Distrito", "Serie 7 días (incidentes/día)", "Δ"]]
        for r in ranking[:12]:
            spark = r.get("sparkline") or []
            serie = " · ".join(str(int(v)) for v in spark) if spark else "—"
            tend = int(r.get("tendencia") or 0)
            spark_table_rows.append(
                [
                    (r.get("distrito") or "—")[:36],
                    serie,
                    f"{'+' if tend > 0 else ''}{tend}",
                ]
            )
        spark_rows_flow.append(
            _styled_data_table(spark_table_rows, [8 * cm, 12 * cm, 2.5 * cm])
        )
        spark_rows_flow.append(Spacer(1, 0.2 * cm))
        spark_rows_flow.append(
            _analysis_box(
                "La serie diaria ayuda a anticipar picos: si el final de la semana supera el inicio "
                "(Δ positivo), refuerce turnos en ese distrito de inmediato. Δ negativo indica "
                "alivio relativo; puede reasignar personal temporalmente hacia distritos en alza.",
                analysis_style,
            )
        )
        spark_rows_flow.append(Spacer(1, 0.35 * cm))
        spark_rows_flow.append(HRFlowable(width="100%", thickness=0.6, color=LINE, spaceAfter=4))
        spark_rows_flow.append(
            Paragraph(
                "Documento generado desde la pestaña Ranking Distritos · CrimeTrack · Uso interno institucional",
                small,
            )
        )
        body.append(KeepTogether(spark_rows_flow))
    else:
        body.append(Spacer(1, 0.35 * cm))
        body.append(HRFlowable(width="100%", thickness=0.6, color=LINE, spaceAfter=4))
        body.append(
            Paragraph(
                "Documento generado desde la pestaña Ranking Distritos · CrimeTrack · Uso interno institucional",
                small,
            )
        )

    doc.build(body)
    return buffer.getvalue()


def build_personal_disponibilidad_pdf(data: dict[str, Any], *, emisor: str) -> bytes:
    """Informe PDF de novedades / disponibilidad del personal de la zona."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=1.4 * cm,
        rightMargin=1.4 * cm,
        topMargin=1.4 * cm,
        bottomMargin=1.4 * cm,
    )
    styles = getSampleStyleSheet()
    title = ParagraphStyle(
        "PersTitle",
        parent=styles["Heading1"],
        fontSize=15,
        textColor=colors.HexColor("#2f4d8a"),
        spaceAfter=6,
    )
    h2 = ParagraphStyle(
        "PersH2",
        parent=styles["Heading2"],
        fontSize=11,
        textColor=colors.HexColor("#334155"),
        spaceBefore=10,
        spaceAfter=6,
    )
    small = ParagraphStyle(
        "PersSmall",
        parent=styles["Normal"],
        fontSize=8,
        textColor=colors.HexColor("#64748b"),
    )
    cell = ParagraphStyle("PersCell", parent=styles["Normal"], fontSize=8, leading=10)

    jur = data.get("jurisdiccion") or {}
    resumen = data.get("resumen") or {}
    personal = data.get("personal") or []
    fecha = data.get("fecha") or ""

    body = [
        Paragraph("Informe de disponibilidad — Personal regional", title),
        Paragraph(
            f"<b>Jurisdicción:</b> {jur.get('nombre') or '—'} ({jur.get('codigo') or '—'})",
            styles["Normal"],
        ),
        Paragraph(f"<b>Fecha de corte:</b> {fecha}", styles["Normal"]),
        Paragraph(f"<b>Elaborado por:</b> {emisor}", styles["Normal"]),
        Paragraph(
            f"<b>Disponibles hoy:</b> {data.get('disponibles_hoy', 0)} / {data.get('total', 0)}",
            styles["Normal"],
        ),
        Spacer(1, 0.35 * cm),
        Paragraph("Resumen por estado operativo", h2),
    ]

    kpi_rows = [
        ["Activos", "Franco", "Vacaciones", "Calamidad", "Arresto", "Permiso"],
        [
            str(resumen.get("ACTIVO", 0)),
            str(resumen.get("FRANCO", 0)),
            str(resumen.get("VACACIONES", 0)),
            str(resumen.get("CALAMIDAD", 0)),
            str(resumen.get("ARRESTO", 0)),
            str(resumen.get("PERMISO", 0)),
        ],
    ]
    kt = Table(kpi_rows, colWidths=[2.8 * cm] * 6)
    kt.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2f4d8a")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#d1d5db")),
                ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#f8fafc")),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    body.append(kt)
    body.append(Spacer(1, 0.4 * cm))
    body.append(Paragraph("Detalle del personal", h2))

    table_data = [
        [
            Paragraph("<b>Nombre</b>", cell),
            Paragraph("<b>Rol</b>", cell),
            Paragraph("<b>Unidad</b>", cell),
            Paragraph("<b>Estado</b>", cell),
            Paragraph("<b>Detalle</b>", cell),
        ]
    ]
    for p in personal:
        table_data.append(
            [
                Paragraph(
                    f"{p.get('nombre') or '—'}<br/><font size='7' color='#64748b'>{p.get('email') or ''}</font>",
                    cell,
                ),
                Paragraph(p.get("rol_label") or p.get("rol") or "—", cell),
                Paragraph(p.get("unidad") or p.get("jurisdiccion") or "—", cell),
                Paragraph(p.get("estado") or "—", cell),
                Paragraph(p.get("estado_detalle") or "—", cell),
            ]
        )
    if len(table_data) == 1:
        table_data.append(
            [Paragraph("Sin personal registrado en la zona.", cell), "", "", "", ""]
        )

    pt = Table(table_data, colWidths=[4.2 * cm, 3.2 * cm, 3.4 * cm, 2.2 * cm, 4.2 * cm])
    pt.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2f4d8a")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#e2e8f0")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    body.append(pt)
    body.append(Spacer(1, 0.45 * cm))
    body.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#cbd5e1")))
    body.append(
        Paragraph(
            "Documento generado desde Gestión de Personal Regional · CrimeTrack · Uso interno institucional",
            small,
        )
    )
    doc.build(body)
    return buffer.getvalue()


def _estado_partes_bar_chart(por_estado: list[dict]) -> Drawing | None:
    rows = [r for r in (por_estado or []) if int(r.get("total") or 0) > 0][:6]
    if not rows:
        return None
    drawing = Drawing(480, 160)
    chart = HorizontalBarChart()
    chart.x = 140
    chart.y = 20
    chart.height = 120
    chart.width = 300
    chart.data = [[int(r.get("total") or 0) for r in rows]]
    chart.categoryAxis.categoryNames = [
        (r.get("label") or r.get("estado") or "-")[:28] for r in rows
    ]
    chart.bars[0].fillColor = PURPLE
    chart.valueAxis.valueMin = 0
    chart.categoryAxis.labels.fontSize = 8
    chart.valueAxis.labels.fontSize = 8
    drawing.add(chart)
    return drawing


def _analisis_estado_partes(data: dict) -> str:
    tasa = data.get("tasa_resolucion") or 0
    aprobado = data.get("aprobado") or 0
    pendiente = data.get("pendiente") or 0
    observado = data.get("observado") or 0
    total = data.get("total") or 0
    partes = [
        f"En el periodo hay <b>{total}</b> partes en la zona. "
        f"La <b>tasa de resolucion</b> es <b>{tasa}%</b> "
        f"({aprobado} aprobados sobre el flujo de control de calidad)."
    ]
    if pendiente:
        partes.append(
            f"Quedan <b>{pendiente}</b> pendientes de revision: priorice la bandeja "
            "de supervisores para no acumular backlog."
        )
    if observado:
        partes.append(
            f"Hay <b>{observado}</b> partes devueltos/observados: revise calidad del "
            "relato e indicios antes de reenviar."
        )
    if tasa >= 80:
        partes.append("La zona mantiene un buen ritmo de cierre de partes.")
    elif tasa < 50 and (aprobado + pendiente + observado) > 0:
        partes.append(
            "La tasa es baja: refuerce supervision y estandarice el llenado del parte."
        )
    return " ".join(partes)


def build_estado_partes_pdf(
    *,
    emisor: str,
    jurisdiccion: dict[str, Any],
    filtros: dict[str, Any],
    estado_partes: dict[str, Any] | None = None,
) -> bytes:
    """PDF de Estado de Partes / Tasa de Resolucion."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=1.2 * cm,
        bottomMargin=1.2 * cm,
    )
    styles = getSampleStyleSheet()
    title = ParagraphStyle(
        "EpTitle",
        parent=styles["Heading1"],
        fontSize=16,
        textColor=PURPLE_DARK,
        spaceAfter=4,
    )
    kicker = ParagraphStyle(
        "EpKicker",
        parent=styles["Normal"],
        fontSize=9,
        textColor=PURPLE,
        fontName="Helvetica-Bold",
        spaceAfter=2,
    )
    h2 = ParagraphStyle(
        "EpH2",
        parent=styles["Heading2"],
        fontSize=12,
        textColor=SLATE,
        spaceBefore=8,
        spaceAfter=6,
    )
    analysis_style = ParagraphStyle(
        "EpAnalysis",
        parent=styles["Normal"],
        fontSize=8.5,
        textColor=SLATE,
        leading=11.5,
    )
    small = ParagraphStyle(
        "EpSmall",
        parent=styles["Normal"],
        fontSize=8,
        textColor=MUTED,
        leading=10,
    )

    data = estado_partes or {}
    generado = datetime.now().strftime("%Y-%m-%d %H:%M")
    body: list[Any] = [
        Paragraph("INTELIGENCIA TACTICA · ESTADO DE PARTES", kicker),
        Paragraph(
            f"Tasa de resolucion — {jurisdiccion.get('nombre') or 'Zona'}",
            title,
        ),
        Paragraph(
            f"Emisor: {emisor} · Generado: {generado} · "
            f"Filtro {filtros.get('fecha_desde') or '-'} -> {filtros.get('fecha_hasta') or '-'}"
            + (
                f" · Distrito: {filtros.get('distrito')}"
                if filtros.get("distrito")
                else ""
            ),
            small,
        ),
        Spacer(1, 0.35 * cm),
        HRFlowable(width="100%", thickness=0.6, color=LINE),
        Spacer(1, 0.35 * cm),
    ]

    kpi_rows = [
        ["Tasa de resolucion", f"{data.get('tasa_resolucion', 0)}%"],
        ["Total partes", str(data.get("total") or 0)],
        ["Aprobados", str(data.get("aprobado") or 0)],
        ["Pendientes", str(data.get("pendiente") or 0)],
        ["Devueltos", str(data.get("observado") or 0)],
        ["Borradores", str(data.get("borrador") or 0)],
    ]
    body.append(_kpi_card_table([[a, b] for a, b in kpi_rows]))
    body.append(Spacer(1, 0.35 * cm))
    body.append(
        Paragraph(
            data.get("nota")
            or "Tasa = Aprobados / (Aprobados + Pendientes + Devueltos).",
            small,
        )
    )

    body.append(Paragraph("Distribucion por estado", h2))
    chart = _estado_partes_bar_chart(data.get("por_estado") or [])
    if chart:
        body.append(chart)
        body.append(Spacer(1, 0.25 * cm))

    table_rows = [["Estado", "Cantidad", "%"]]
    for r in data.get("por_estado") or []:
        table_rows.append(
            [
                r.get("label") or r.get("estado") or "-",
                str(r.get("total") or 0),
                f"{r.get('pct') or 0}%",
            ]
        )
    if len(table_rows) > 1:
        body.append(_styled_data_table(table_rows, [8 * cm, 3.5 * cm, 3 * cm]))

    body.append(Spacer(1, 0.4 * cm))
    body.append(Paragraph("Lectura tactica", h2))
    body.append(_analysis_box(_analisis_estado_partes(data), analysis_style))
    body.append(Spacer(1, 0.45 * cm))
    body.append(HRFlowable(width="100%", thickness=0.5, color=LINE))
    body.append(
        Paragraph(
            "Documento generado desde Dashboard de zona · CrimeTrack · Uso interno",
            small,
        )
    )
    doc.build(body)
    return buffer.getvalue()
